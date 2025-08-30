from bs4 import BeautifulSoup
import requests
import pandas as pd
import time
from openpyxl.styles import Alignment,Font
from openpyxl import load_workbook


url="https://internshala.com/jobs/"

headers = {
    'Accept': 'application/x-clarity-gzip',
    'Accept-Encoding': 'gzip, deflate, br, zstd',
    'Accept-Language': 'en-US,en;q=0.9,ml;q=0.8,ja;q=0.7',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/139.0.0.0 Safari/537.36'
    }

response=requests.get(url)
soup=BeautifulSoup(response.text,"html.parser")

cards = soup.find_all("div", class_="internship_meta experience_meta")
jobs=[]

for card in cards:
    try:
        job_name = card.find("a", class_="job-title-href").text.strip()
        job_url = "https://internshala.com" + card.find("a", class_="job-title-href")["href"]
    except:
        continue
    location=""
    Exp=""
    skills=""
    salary=""
    about=""

    if job_url:
        try:
            detail_resp = requests.get(job_url, headers=headers)
            detail_soup = BeautifulSoup(detail_resp.text, "html.parser")
            
            location_element = detail_soup.find("p", id="location_names")
            if location_element:
                location = location_element.find("a").text.strip() if location_element.find("a") else location_element.text.strip()

            experience_element = detail_soup.find("div", class_="job-experience-item")
            if experience_element:
                Exp = experience_element.find("div", class_="item_body").text.strip()

            skills_div = detail_soup.find("div", class_="round_tabs_container")
            if skills_div:
                skill_spans = skills_div.find_all("span", class_="round_tabs")
                skills = ", ".join([s.text.strip() for s in skill_spans])

            internship_details = detail_soup.find("div", class_="internship_details")
            if internship_details:
                text_container = internship_details.find("div", class_="text-container")
                if text_container:
                    lines = [line.strip() for line in text_container.text.split('\n') if line.strip()]
                    about = lines[0:10]
            
            salary_container = detail_soup.find("div", class_="text-container salary_container")
            if salary_container:
                salary = salary_container.p.text.strip()
            
        except Exception as e:
            print(f"Error scraping details from {job_url}: {e}")
            skills = ""
            about = ""

        time.sleep(0.5)

    jobs.append({
        "JobTitle": job_name,   
        "Location": location,
        "Experience": Exp,
        "Skills": skills,
        "Salary": salary,
        "JobUrl": job_url,
        "JobDescriptionSummary": about
    })

df = pd.DataFrame(jobs)

file = "Jobs.xlsx"
df.to_excel(file, index=False, engine="openpyxl")   

wb = load_workbook(file)
sheet = wb.active
sheet.title = "Jobs"

# Add headers
wb = load_workbook(file)
sheet = wb.active

# Auto-adjust column widths
for column_cells in sheet.columns:
    length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
    sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

wb.save(file)
print("Completed")
